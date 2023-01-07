import requests
import json
import openpyxl
from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtWidgets import QApplication, QWidget, QMainWindow, QDesktopWidget, QLabel, QPushButton, QComboBox, QFileDialog, QMessageBox, QVBoxLayout
def insert_coordinate(sheet_name):
    data = openpyxl.load_workbook(sheet_name)
    sheet = data.active
    missing_num = 0
    dataset1 = openpyxl.load_workbook("dataset.xlsx")
    anssheet = dataset1.active
    sheet.cell(1, 11, 'latitude')
    sheet.cell(1, 12, 'longitude')
    maxrow_sheet = sheet.max_row
    maxrow_ans = anssheet.max_row
    address_col = 0
    a = True
    latitude_col = 0
    la = True
    longitude_col = 0
    lo = True
    name_col = 0
    n = True
    for i in range(1, sheet.max_column + 1):
        if(a == True and sheet.cell(column = i, row = 1).value == 'address'):
            address_col = i
            a = False
        if(la == True and sheet.cell(column = i, row = 1).value == 'latitude'):
            latitude_col = i
            la = False
        if(lo == True and sheet.cell(column = i, row = 1).value == 'longitude'):
            longitude_col = i
            lo = False
        if(n == True and sheet.cell(column = i, row = 1).value == 'name'):
            name_col = i
            n = False
        QApplication.processEvents()
    for i in range (2, maxrow_sheet+1):
        address = sheet.cell(column = address_col, row = i).value
        repeat = False
        if(sheet.cell(column=latitude_col, row=i).value != None and sheet.cell(column=longitude_col,row=i).value != None):
            continue
        for j in range (2, maxrow_ans+1):
            if(address == anssheet.cell(column = 4, row = j).value):
                repeat = True
                sheet.cell(i, 11, anssheet.cell(column = 2, row = j).value)
                sheet.cell(i, 12, anssheet.cell(column = 3, row = j).value)
                QApplication.processEvents()
                break
        if(repeat == False):
            

            par = {'address': address, 'key': '7bd0aec565f20d16cca55242f46a47ee'}
            url = 'http://restapi.amap.com/v3/geocode/geo'
            try:
                res = requests.get(url, par)
            except:
                i = i - 1
                continue
            json_data = json.loads(res.text)
            try:
                geo = json_data['geocodes'][0]['location']
            except:
                missing_num = missing_num + 1
                continue
            name = sheet.cell(column = name_col, row = i).value        
            latitude = geo.split(',')[1]
            longitude = geo.split(',')[0]
            anssheet.cell(maxrow_ans+1, 2, latitude)
            anssheet.cell(maxrow_ans+1, 3, longitude)
            anssheet.cell(maxrow_ans+1, 1, name)
            anssheet.cell(maxrow_ans+1, 4, address)
            sheet.cell(i, 11, latitude)
            sheet.cell(i, 12, longitude)
            maxrow_ans = maxrow_ans + 1
            QApplication.processEvents()
            
    dataset1.save('dataset.xlsx')        
    data.save(sheet_name)



    
