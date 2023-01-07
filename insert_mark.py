import folium
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from folium.plugins import HeatMap
from folium import plugins
from folium import FeatureGroup, LayerControl, Map, Marker
from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtWidgets import QApplication, QWidget, QMainWindow, QDesktopWidget, QLabel, QPushButton, QComboBox, QFileDialog, QMessageBox, QVBoxLayout
m = folium.Map()
Customer_yes = 0
Engineer_yes = 0

def insert_customer(sheet_name):
    
    global Customer_yes
    Customer_yes = 1
    data = openpyxl.load_workbook("Customer_temp.xlsx")
    sheet = data.active
    i = sheet.max_row + 1
    dataset = pd.read_excel(sheet_name)
    tooltip = "Click me!"
    templat = " "
    temptru = " "
    templng = " "
    for lat, lng, name, truck in zip(dataset.latitude, dataset.longitude, dataset.name, dataset.truck):
        if(lat == templat and lng == templng and truck == temptru):
            QApplication.processEvents()
            continue
        templat = lat
        templng = lng
        temptru = truck
        sheet.cell(i, 1, lat)
        sheet.cell(i, 2, lng)
        sheet.cell(i, 3, name)
        sheet.cell(i, 4, truck)
        i = i + 1
        QApplication.processEvents()
    data.save("Customer_temp.xlsx")
    
def show_customer():
    
    incidents_customer = plugins.MarkerCluster()
    tempdata = pd.read_excel("Customer_temp.xlsx")
    for lat, lng, name, truck in zip(tempdata.latitude, tempdata.longitude, tempdata.name, tempdata.truck):
        folium.Marker(
            [lat, lng], popup=name, tooltip="Click me!", icon=folium.Icon(color="gray", icon="info-sign")
        ).add_to(incidents_customer)
        QApplication.processEvents()
    incidents_customer.add_to(folium.FeatureGroup(name="Customer").add_to(m))
    
def insert_engineer(sheet_name):
    
    global Engineer_yes
    Engineer_yes = 1
    dataset = pd.read_excel(sheet_name)
    data = openpyxl.load_workbook("Engineer_temp.xlsx")
    sheet = data.active
    i = sheet.max_row + 1
    for lat, lng, name in zip(dataset.latitude, dataset.longitude, dataset.name):
        sheet.cell(i, 1, lat)
        sheet.cell(i, 2, lng)
        sheet.cell(i, 3, name)
        i = i + 1
        QApplication.processEvents()
    data.save("Engineer_temp.xlsx")
    
def show_engineer():
    
    tooltip = "Click me!"
    incidents_engineer = plugins.MarkerCluster()
    tempdata = pd.read_excel("Engineer_temp.xlsx")
    for lat, lng, name in zip(tempdata.latitude, tempdata.longitude, tempdata.name):
        folium.Marker(
            [lat, lng], popup=name, tooltip=tooltip
        ).add_to(incidents_engineer)
        QApplication.processEvents()
    incidents_engineer.add_to(folium.FeatureGroup(name="Engineer").add_to(m))
    
def make_heat_map(typel):
    
    if(typel == 'Engineer'):
        url = 'Engineer_temp.xlsx'
    else:
        url = 'Customer_temp.xlsx'
    dataset = pd.read_excel(url)
    data = dataset[['latitude', 'longitude']].values.tolist()
    HeatMap(data).add_to(folium.FeatureGroup(name= typel + ' Heat Map', show = False).add_to(m))
    
def show_heat_map(typel):
    if(typel == 'Engineer'):
        url = 'Engineer_temp.xlsx'
    else:
        url = 'Customer_temp.xlsx'
    dataset = pd.read_excel(url)
    data = dataset[['latitude', 'longitude']].values.tolist()
    HeatMap(data).add_to(folium.FeatureGroup(name= typel + ' Heat Map', show = False).add_to(m))
def delete(sheet):
    
    sheet.delete_rows(2, sheet.max_row-1)       
    return
