import requests 
import json
import openpyxl
from requests import get
latitude = 0
longitude = 0
def iptolocation():
    global latitude, longitude
    data = openpyxl.load_workbook("index.xlsx")
    ip = get('http://api.ipify.org').text
    sheet = data.active
    if(sheet.cell(column=3, row=1).value != None and ip == sheet.cell(column=3,row=1).value):
        latitude = sheet.cell(column=1,row=1).value
        longitude = sheet.cell(column=2,row=1).value
        return

    par = {'type': '4','ip': ip, 'key': '7bd0aec565f20d16cca55242f46a47ee'}
    url = 'http://restapi.amap.com/v5/ip'
    res = requests.get(url, par)
    json_data = json.loads(res.text)
    ip_address = json_data['location']
    latitude = ip_address.split(',')[1]
    longitude = ip_address.split(',')[0]
    sheet.cell(1,1,latitude)
    sheet.cell(1,2,longitude)
    sheet.cell(1,3,ip)
    data.save("index.xlsx")    
