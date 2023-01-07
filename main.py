import os
import sys
import json
import folium
import os.path
import openpyxl
import requests
import webbrowser
import insert_mark
import pandas as pd
import importlib.util
import address_search
import ip_geolocation
from folium import plugins
from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtWidgets import QApplication, QWidget, QMainWindow, QDesktopWidget, QLabel, QPushButton, QComboBox, QFileDialog, QMessageBox, QVBoxLayout
from openpyxl import load_workbook
from folium.plugins import HeatMap
from PyQt5.QtWebEngineWidgets import QWebEngineView
from folium import FeatureGroup, LayerControl, Map, Marker

fn_engineer = ""
fn_customer = ""
Missing = 0
content = 'AMap (Default)'
        
class Loading(QWidget):
    def __init__(self):
        
        super(Loading, self).__init__()
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('Loading!')
        self.btn = QPushButton('Start Loading')
        self.setWindowIcon(QIcon('touch-icon-png-data.jpg'))
        
        self.btn.clicked.connect(self.doit)
        self.btn.clicked.connect(self.mark)
        self.resize(400, 160)
        self.vbox = QVBoxLayout()
        self.vbox.addWidget(self.btn)
        self.setLayout(self.vbox)
        
    def doit(self):
        
        ip_geolocation.iptolocation()
        if(content == "Google Map (VPN needed)"):
            insert_mark.m = folium.Map(location=[ip_geolocation.latitude, ip_geolocation.longitude],tiles='https://mt.google.com/vt/lyrs=m&x={x}&y={y}&z={z}',attr='default', zoom_start=5)
        elif(content == 'AMap (Default)'):
            insert_mark.m = folium.Map(location=[ip_geolocation.latitude, ip_geolocation.longitude],tiles='http://webrd02.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=1&style=7&x={x}&y={y}&z={z}',attr='default', zoom_start=5)

        if(Missing == 0 or Missing == 2):
            address_search.insert_coordinate(fn_engineer[0])
            insert_mark.insert_engineer(fn_engineer[0])
        if(Missing == 0 or Missing == 1):
            address_search.insert_coordinate(fn_customer[0])
            insert_mark.insert_customer(fn_customer[0])
    def mark(self):
        
        global fn_customer
        global fn_engineer
        insert_mark.show_engineer()
        insert_mark.show_heat_map('Engineer')
        insert_mark.show_customer()
        insert_mark.show_heat_map('Customer')
        self.close()
        fn_customer=""
        fn_engineer=""
        LayerControl().add_to(insert_mark.m)
        url = 'index.html'
        insert_mark.m.save(url)
        webbrowser.open(url, new=2)
        


class Main(QMainWindow):
    def __init__(self):
        super(Main, self).__init__()
        
        self.initUI()
        
    def initUI(self):
        self.setGeometry(600, 250, 750, 350)
        self.statusbar = self.statusBar()
        self.center()

        self.setWindowTitle('Zehao Song')

        self.setWindowIcon(QIcon('touch-icon-png-data.jpg'))

        self.UiComponents()
        txt_topic = QLabel("Map Generator :)", self)
        txt_topic.setGeometry(30, 10, 400, 75)
        txt_topic.setFont(QFont('Terminus', 28, QFont.Bold))
        
        one = QLabel("↓1.",self)
        one.setGeometry(5,90,30,30)
        one.setFont(QFont('Times', 15))
        
        reload_btn = QPushButton("Reload Excel", self)
        reload_btn.setGeometry(185, 90, 100, 45)
        reload_btn.setFont(QFont("Times", 8))
        reload_btn.setToolTip('Clean your memory')
        reload_btn.clicked.connect(self.reload_excel)

        current_btn = QPushButton("Show Current", self)
        current_btn.setGeometry(45, 90, 100, 45)
        current_btn.setFont(QFont("Times", 8))
        current_btn.clicked.connect(self.current)

        two = QLabel("↓2.",self)
        two.setGeometry(5,160,30,30)
        two.setFont(QFont('Times', 15))

        thr = QLabel("↓3.",self)
        thr.setGeometry(5,240,30,30)
        thr.setFont(QFont('Times', 15))
        
        excel_btn1 = QPushButton("Upload Engineer", self)
        excel_btn1.setGeometry(45, 225, 150, 50)
        excel_btn1.setFont(QFont('Times', 10))
        excel_btn1.clicked.connect(self.engineer_open)
        
        excel_btn2 = QPushButton("Upload Customer", self)
        excel_btn2.setGeometry(250, 225, 150, 50)
        excel_btn2.setFont(QFont('Times', 10))
        excel_btn2.clicked.connect(self.customer_open)

        continue_btn = QPushButton("Continue", self)
        continue_btn.setGeometry(450, 225, 150, 50)
        continue_btn.setFont(QFont("Times", 10))
        continue_btn.clicked.connect(self.Continue)

    def current(self):
        
        webbrowser.open("index.html", new=2)
        
    def UiComponents(self):
  
        self.combo_box = QComboBox(self)
  
        self.combo_box.setGeometry(45, 160, 175, 30)
  
        map_list = ['AMap (Default)',"Google Map (VPN needed)"]
  
        self.combo_box.addItems(map_list)
  
        button = QPushButton("Select Map ", self)
        button.setGeometry(250, 160, 175, 30)
        button.pressed.connect(self.find)
        
    def reload_excel(self):

        book = openpyxl.load_workbook('Engineer_temp.xlsx')
        insert_mark.delete(book['Sheet1'])
        book.save('Engineer_temp.xlsx')
        book = openpyxl.load_workbook('Customer_temp.xlsx')
        insert_mark.delete(book['Sheet1'])
        book.save('Customer_temp.xlsx')
        self.statusbar.showMessage("Excel is reloaded")
        insert_mark.m = folium.Map(location=[ip_geolocation.latitude, ip_geolocation.longitude],tiles='https://mt.google.com/vt/lyrs=m&x={x}&y={y}&z={z}',attr='default', zoom_start=5)
        insert_mark.m.save("index.html")
        
    def find(self):
        
        global content
        content = self.combo_box.currentText()
        self.statusbar.showMessage(content + " is chosen")

    def center(self):

        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
        
    def engineer_open(self):
        global fn_engineer
        path = QFileDialog.getOpenFileName(self, 'Open a file', '',
                                        'All Files (*.*)')
        if(path != ('', '') and path[0].lower().endswith('.xlsx')):
            fn_engineer = path
            self.statusbar.showMessage(fn_engineer[0] + " is uploaded")
        else:
            msg = QMessageBox()
            msg.setWindowTitle("Warning!")
            msg.setText("Must be Excel Type!")
            msg.setIcon(QMessageBox.Warning)
            x = msg.exec_()
            
    def customer_open(self):
        global fn_customer
        path = QFileDialog.getOpenFileName(self, 'Open a file', '',
                                        'All Files (*.*)')
        if(path != ('', '') and path[0].lower().endswith('.xlsx')):
            fn_customer = path
            self.statusbar.showMessage(fn_customer[0] + " is uploaded")
        else:
            msg = QMessageBox()
            msg.setWindowTitle("Warning!")
            msg.setText("Must be Excel Type!")
            msg.setIcon(QMessageBox.Warning)
            x = msg.exec_()
            
    def Continue(self):
        global Missing
        if(fn_engineer == "" and fn_customer == ""):
            msg = QMessageBox()
            msg.setWindowTitle("Warning!")
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Both excel is missing!")
            x = msg.exec_()
            return
        elif(fn_engineer == ""):
            Missing = 1
            msg = QMessageBox()
            msg.setWindowTitle("Warning!")
            msg.setIcon(QMessageBox.Information)
            msg.setText("Engineer excel is missing!")
            x = msg.exec_()
        elif(fn_customer == ""):
            Missing = 2
            msg = QMessageBox()
            msg.setWindowTitle("Warning!")
            msg.setIcon(QMessageBox.Information)
            msg.setText("Customer excel is missing!")
            x = msg.exec_()
        else:
            Missing = 0
        self.w = Loading()
        self.w.show()


            
if __name__ == '__main__':
    #创建应用程序和对象
    app = QApplication(sys.argv)
    ex = Main()
    ex.show()
    sys.exit(app.exec_()) 
